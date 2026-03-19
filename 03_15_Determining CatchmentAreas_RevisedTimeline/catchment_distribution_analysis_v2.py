"""
catchment_distribution_analysis_v2.py
--------------------------------------
For each of three hospital / time-window combinations, computes:
  - Which locations from population_density_infrasturtcure.xlsx fall inside
    the hospital catchment area
  - A population density distribution count table (rows=category, cols=months)
  - An infrastructure level distribution count table (rows=category, cols=months)

Hospitals & windows:
  1. Al Shifa Medical Hospital  — 10/07/2023 to 11/03/2023
  2. Nasser Hospital            — 11/11/2024 to 02/01/2025
  3. European Hospital (EGH)    — 12/11/2023 to 04/28/2024
     (EGH has 4 Voronoi sub-periods; the majority-days rule assigns each
      calendar month to the sub-period catchment that covers most of it)

Output: catchment_distribution_analysis_v2.xlsx
  9 sheets total per hospital (3 hospitals × 3 sheets):
    {Key}_Locations       — list of in-catchment locations with lat/lon
    {Key}_PopDensity      — monthly distribution count table
    {Key}_Infrastructure  — monthly distribution count table

Terminal output: prints the location list and both tables for all hospitals.

Place this script in the same folder as:
  - population_density_infrasturtcure.xlsx
  - pse_admin2.geojson
"""

from calendar import monthrange
from datetime import datetime
from pathlib import Path
from typing import Any, Callable, Dict, List, Tuple

import numpy as np
import pandas as pd
import geopandas as gpd
from shapely.geometry import Point, Polygon, MultiPolygon, MultiPoint, box
from shapely.ops import unary_union
from pyproj import Geod
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

try:
    from shapely.ops import voronoi_diagram
    from shapely import geometry as shapely_geom
    _HAS_SHAPELY_VORONOI = True
except ImportError:
    voronoi_diagram = None
    shapely_geom = None
    _HAS_SHAPELY_VORONOI = False

from scipy.spatial import Voronoi

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------
BASE_PATH    = Path(__file__).parent.resolve()
DENSITY_PATH = BASE_PATH / "population_density_infrasturtcure.xlsx"
GAZA_GEOJSON = BASE_PATH / "pse_admin2.geojson"
OUTPUT_PATH  = BASE_PATH / "catchment_distribution_analysis_v2.xlsx"

CRS_WGS84             = "EPSG:4326"
CRS_WEBMERC           = "EPSG:3857"
CATCHMENT_DISTANCE_KM = 5.0
VORONOI_BUFFER_METERS = 15_000

# ---------------------------------------------------------------------------
# Category ordering
# ---------------------------------------------------------------------------
POP_ORDER   = ["High", "Medium", "Low"]
INFRA_ORDER = ["Urban", "Camp", "Suburban", "Rubble"]

# ---------------------------------------------------------------------------
# Hospital configurations
# ---------------------------------------------------------------------------
HOSPITAL_CONFIGS: List[Dict] = [
    {
        "key"   : "AlShifa",
        "name"  : "Al Shifa Medical Hospital",
        "window": (datetime(2023, 10,  7), datetime(2023, 11,  3)),
        "sub_periods": [
            {
                "label": "Full window",
                "start": datetime(2023, 10,  7),
                "end"  : datetime(2023, 11,  3),
                "open" : [
                    ("Al Shifa Medical Hospital",  31.52369093, 34.44274363),
                    ("Al-Quds Hospital",           31.5056094,  34.4297735 ),
                    ("Nasser Hospital",            31.34689036, 34.29193795),
                    ("European Hospital",          31.303174,   34.31925517),
                    ("Kuwait Hospital",            31.28812189, 34.24915904),
                    ("Al-Aqsa Hospital",           31.419969,   34.36      ),
                ],
            },
        ],
    },
    {
        "key"   : "Nasser",
        "name"  : "Nasser Hospital",
        "window": (datetime(2024, 11, 11), datetime(2025,  2,  1)),
        "sub_periods": [
            {
                "label": "Full window",
                "start": datetime(2024, 11, 11),
                "end"  : datetime(2025,  2,  1),
                "open" : [
                    ("Al Shifa Medical Hospital",  31.52369093, 34.44274363),
                    ("Nasser Hospital",            31.34689036, 34.29193795),
                    ("European Hospital",          31.303174,   34.31925517),
                    ("Kuwait Hospital",            31.28812189, 34.24915904),
                ],
            },
        ],
    },
    {
        "key"   : "EGH",
        "name"  : "European Hospital",
        "window": (datetime(2023, 12, 11), datetime(2024,  4, 28)),
        "sub_periods": [
            {
                "label": "SP1: 12/11/2023–02/19/2024",
                "start": datetime(2023, 12, 11),
                "end"  : datetime(2024,  2, 19),
                "open" : [
                    ("Al Shifa Medical Hospital",  31.52369093, 34.44274363),
                    ("Nasser Hospital",            31.34689036, 34.29193795),
                    ("European Hospital",          31.303174,   34.31925517),
                    ("Kuwait Hospital",            31.28812189, 34.24915904),
                    ("Al-Aqsa Hospital",           31.419969,   34.36      ),
                ],
            },
            {
                "label": "SP2: 02/20/2024–03/17/2024",
                "start": datetime(2024,  2, 20),
                "end"  : datetime(2024,  3, 17),
                "open" : [
                    ("Al Shifa Medical Hospital",  31.52369093, 34.44274363),
                    ("European Hospital",          31.303174,   34.31925517),
                    ("Kuwait Hospital",            31.28812189, 34.24915904),
                    ("Al-Aqsa Hospital",           31.419969,   34.36      ),
                ],
            },
            {
                "label": "SP3: 03/18/2024–03/31/2024",
                "start": datetime(2024,  3, 18),
                "end"  : datetime(2024,  3, 31),
                "open" : [
                    ("European Hospital",          31.303174,   34.31925517),
                    ("Kuwait Hospital",            31.28812189, 34.24915904),
                    ("Al-Aqsa Hospital",           31.419969,   34.36      ),
                ],
            },
            {
                "label": "SP4: 04/01/2024–04/28/2024",
                "start": datetime(2024,  4,  1),
                "end"  : datetime(2024,  4, 28),
                "open" : [
                    ("Al Shifa Medical Hospital",  31.52369093, 34.44274363),
                    ("European Hospital",          31.303174,   34.31925517),
                    ("Kuwait Hospital",            31.28812189, 34.24915904),
                    ("Al-Aqsa Hospital",           31.419969,   34.36      ),
                ],
            },
        ],
    },
]

# ---------------------------------------------------------------------------
# Gaza boundary
# ---------------------------------------------------------------------------

def load_gaza_boundary(path: Path) -> Any:
    gdf = gpd.read_file(path)
    if gdf.crs is None:
        gdf = gdf.set_crs(CRS_WGS84)
    return gdf.to_crs(CRS_WGS84).unary_union


# ---------------------------------------------------------------------------
# Voronoi helpers
# ---------------------------------------------------------------------------

def _extract_voronoi_polygons(vor, bbox_proj, hosp_proj, hosp_gdf) -> Dict[str, Any]:
    poly_list = []
    if hasattr(vor, "geoms"):
        for g in vor.geoms:
            if isinstance(g, (Polygon, MultiPolygon)):
                poly_list.append(g)
    elif isinstance(vor, (Polygon, MultiPolygon)):
        poly_list = [vor]
    if not poly_list:
        return {}
    polys_proj    = gpd.GeoDataFrame(geometry=poly_list, crs=CRS_WEBMERC)
    hosp_pts_proj = hosp_proj.geometry
    assigned: List[Tuple[str, Any]] = []
    for poly in polys_proj.geometry:
        if poly is None or poly.is_empty:
            continue
        rep         = poly.representative_point()
        dists       = hosp_pts_proj.distance(rep)
        nearest_idx = int(dists.idxmin())
        hosp_name   = hosp_gdf.iloc[nearest_idx]["Hospital"]
        poly_wgs    = gpd.GeoSeries([poly], crs=CRS_WEBMERC).to_crs(CRS_WGS84).iloc[0]
        assigned.append((hosp_name, poly_wgs))
    result: Dict[str, Any] = {}
    for hosp in hosp_gdf["Hospital"].values:
        polys_for    = [p for (h, p) in assigned if h == hosp]
        result[hosp] = unary_union(polys_for) if polys_for else Polygon()
    return result


def _scipy_voronoi_clipped(hosp_proj, hosp_gdf, bbox_proj, gaza_union) -> Dict[str, Any]:
    coords     = np.array([(pt.x, pt.y) for pt in hosp_proj.geometry])
    vor        = Voronoi(coords)
    hosp_names = hosp_gdf["Hospital"].values

    def make_bounded_poly(site):
        bbox_coords = np.array(bbox_proj.exterior.coords)
        pts         = np.vstack([site, bbox_coords])
        mp          = MultiPoint([tuple(p) for p in pts])
        return mp.convex_hull.intersection(bbox_proj)

    proj_polys = []
    for pt_idx, region_idx in enumerate(vor.point_region):
        region = vor.regions[region_idx]
        if not region or -1 in region:
            poly = make_bounded_poly(coords[pt_idx])
        else:
            try:
                verts = [vor.vertices[i] for i in region]
                poly  = Polygon(verts).intersection(bbox_proj)
            except Exception:
                poly  = make_bounded_poly(coords[pt_idx])
        proj_polys.append(poly)

    polys_wgs = gpd.GeoDataFrame(geometry=proj_polys, crs=CRS_WEBMERC).to_crs(CRS_WGS84)
    result: Dict[str, Any] = {}
    for i, hosp_name in enumerate(hosp_names):
        try:
            clipped = polys_wgs.iloc[i].geometry.intersection(gaza_union)
        except Exception:
            clipped = Polygon()
        result[hosp_name] = clipped if clipped and not clipped.is_empty else Polygon()

    covered  = unary_union([g for g in result.values() if g and not g.is_empty])
    leftover = gaza_union.difference(covered) if covered else gaza_union
    if leftover and not leftover.is_empty:
        pieces        = [leftover] if isinstance(leftover, Polygon) else list(leftover.geoms)
        hosp_pts_proj = hosp_proj.geometry
        for piece in pieces:
            if piece.is_empty:
                continue
            piece_proj  = gpd.GeoSeries([piece], crs=CRS_WGS84).to_crs(CRS_WEBMERC).iloc[0]
            rep         = piece_proj.representative_point()
            dists       = hosp_pts_proj.distance(rep)
            nearest_idx = int(dists.idxmin())
            n           = hosp_names[nearest_idx]
            cur         = result.get(n, Polygon())
            result[n]   = cur.union(piece) if cur and not cur.is_empty else piece
    return result


def compute_catchment(
    target_hospital: str,
    open_hospitals: List[Tuple[str, float, float]],
    gaza_union: Any,
) -> Polygon:
    hosp_gdf  = gpd.GeoDataFrame(
        {"Hospital": [h[0] for h in open_hospitals]},
        geometry=[Point(h[2], h[1]) for h in open_hospitals],
        crs=CRS_WGS84,
    )
    hosp_proj = hosp_gdf.to_crs(CRS_WEBMERC)
    gaza_proj = gpd.GeoSeries([gaza_union], crs=CRS_WGS84).to_crs(CRS_WEBMERC)
    minx, miny, maxx, maxy = gaza_proj.total_bounds
    bbox_proj = box(
        minx - VORONOI_BUFFER_METERS, miny - VORONOI_BUFFER_METERS,
        maxx + VORONOI_BUFFER_METERS, maxy + VORONOI_BUFFER_METERS,
    )
    if _HAS_SHAPELY_VORONOI:
        try:
            multip    = shapely_geom.MultiPoint([(pt.x, pt.y) for pt in hosp_proj.geometry])
            vor       = voronoi_diagram(multip, envelope=bbox_proj, tolerance=0.0)
            polys_map = _extract_voronoi_polygons(vor, bbox_proj, hosp_proj, hosp_gdf)
        except Exception:
            polys_map = _scipy_voronoi_clipped(hosp_proj, hosp_gdf, bbox_proj, gaza_union)
    else:
        polys_map = _scipy_voronoi_clipped(hosp_proj, hosp_gdf, bbox_proj, gaza_union)

    voronoi_poly = polys_map.get(target_hospital, Polygon())
    clipped      = voronoi_poly.intersection(gaza_union)

    hosp_row    = hosp_gdf[hosp_gdf["Hospital"] == target_hospital]
    pt          = hosp_row.geometry.iloc[0]
    geod        = Geod(ellps="WGS84")
    circle_pts  = [geod.fwd(pt.x, pt.y, a, CATCHMENT_DISTANCE_KM * 1000)[:2]
                   for a in np.linspace(0, 360, 128)]
    circle_poly = Polygon(circle_pts)
    catchment   = clipped.intersection(circle_poly)
    if isinstance(catchment, MultiPolygon):
        catchment = max(catchment.geoms, key=lambda p: p.area)
    return catchment


# ---------------------------------------------------------------------------
# Month utilities
# ---------------------------------------------------------------------------

def months_in_window(start: datetime, end: datetime) -> List[str]:
    """All 'YYYY-MM' strings for calendar months overlapping [start, end]."""
    months, y, m = [], start.year, start.month
    while (y, m) <= (end.year, end.month):
        months.append(f"{y}-{m:02d}")
        m += 1
        if m > 12:
            m, y = 1, y + 1
    return months


def catchment_for_month(
    month_str  : str,
    sub_periods: List[Dict],
    catchments : Dict[int, Polygon],
    window     : Tuple[datetime, datetime],
) -> Polygon:
    """
    For months that span a sub-period boundary, return the catchment polygon
    whose sub-period covers the most days within that month (clipped to window).
    """
    year, mon = int(month_str[:4]), int(month_str[5:])
    _, last_day  = monthrange(year, mon)
    month_start  = max(datetime(year, mon,       1), window[0])
    month_end    = min(datetime(year, mon, last_day), window[1])
    best_idx, best_days = 0, 0
    for idx, sp in enumerate(sub_periods):
        overlap_start = max(sp["start"], month_start)
        overlap_end   = min(sp["end"],   month_end)
        days = (overlap_end - overlap_start).days + 1
        if days > best_days:
            best_days, best_idx = days, idx
    return catchments[best_idx]


# ---------------------------------------------------------------------------
# Location finder: returns DataFrame of in-catchment rows
# ---------------------------------------------------------------------------

def find_locations_in_catchment(
    df        : pd.DataFrame,
    catchment : Polygon,
) -> pd.DataFrame:
    """Return rows of df whose (longitude, latitude) point lies within catchment."""
    mask = df.apply(
        lambda row: Point(row["longitude"], row["latitude"]).within(catchment),
        axis=1,
    )
    return df[mask].copy()


def get_union_catchment_locations(
    df          : pd.DataFrame,
    sub_periods : List[Dict],
    catchments  : Dict[int, Polygon],
) -> pd.DataFrame:
    """
    For hospitals with multiple sub-periods (EGH), return the UNION of all
    locations that appear in ANY sub-period catchment, tagged with which
    sub-period(s) they belong to.
    """
    if len(sub_periods) == 1:
        locs = find_locations_in_catchment(df, catchments[0]).copy()
        locs["sub_periods"] = sub_periods[0]["label"]
        return locs

    all_rows = []
    for idx, sp in enumerate(sub_periods):
        locs = find_locations_in_catchment(df, catchments[idx])
        locs = locs.copy()
        locs["_sp_idx"] = idx
        locs["_sp_label"] = sp["label"]
        all_rows.append(locs)

    combined = pd.concat(all_rows, ignore_index=True)

    # Group by location name; collect which sub-periods each appears in
    def aggregate_sp(group):
        labels = sorted(group["_sp_label"].unique(), key=lambda x: int(x[2]))
        first  = group.iloc[0]
        return pd.Series({
            "location"    : first["location"],
            "latitude"    : first["latitude"],
            "longitude"   : first["longitude"],
            "sub_periods" : " | ".join(labels),
        })

    result = (
        combined.groupby("location", sort=False)
        .apply(aggregate_sp)
        .reset_index(drop=True)
    )
    return result


# ---------------------------------------------------------------------------
# Distribution table builder
# ---------------------------------------------------------------------------

def build_distribution_table(
    catchment_fn  : Callable[[str], Polygon],
    df            : pd.DataFrame,
    window_months : List[str],
    category_order: List[str],
    index_name    : str,
) -> pd.DataFrame:
    available = [m for m in window_months if m in df.columns]
    counts    = {cat: {} for cat in category_order}
    for month_str in available:
        catchment = catchment_fn(month_str)
        in_c      = df.apply(
            lambda row: Point(row["longitude"], row["latitude"]).within(catchment),
            axis=1,
        )
        subset = df[in_c]
        for cat in category_order:
            counts[cat][month_str] = int((subset[month_str] == cat).sum())

    result = pd.DataFrame(counts).T
    result = result.reindex(index=category_order, columns=available, fill_value=0)
    result["TOTAL"]     = result.sum(axis=1)
    result.index.name   = index_name
    return result


# ---------------------------------------------------------------------------
# Excel writer
# ---------------------------------------------------------------------------

def write_excel(hospital_results: List[Dict], output_path: Path) -> None:
    """
    hospital_results: list of dicts with keys:
      key, name, window_str,
      locations_pop  (DataFrame),
      locations_infra(DataFrame),
      pop_table      (DataFrame),
      infra_table    (DataFrame),
      sub_period_info(list of dicts with label/start/end/area_km2)
    """
    wb = Workbook()
    wb.remove(wb.active)

    # Colour palette
    NAVY   = "1F4E79";  DARK   = "0D2B55";  STEEL  = "2E75B6"
    LBLUE  = "D6E4F0";  ALT    = "EBF3FB";  WHITE  = "FFFFFF"
    GREEN  = "375623";  LGREEN = "E2EFDA"

    def fill(hex_col):
        return PatternFill("solid", start_color=hex_col, end_color=hex_col)

    def font(bold=False, color="000000", size=10, italic=False):
        return Font(name="Arial", bold=bold, color=color, size=size, italic=italic)

    thin   = Side(style="thin", color="BFBFBF")
    BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
    CENTRE = Alignment(horizontal="center", vertical="center", wrap_text=True)
    LEFT   = Alignment(horizontal="left",   vertical="center")
    RIGHT  = Alignment(horizontal="right",  vertical="center")

    def hdr_cell(ws, row, col, value):
        c = ws.cell(row=row, column=col, value=value)
        c.font = font(bold=True, color="FFFFFF"); c.fill = fill(NAVY)
        c.alignment = CENTRE; c.border = BORDER
        return c

    def title_cell(ws, row, col, value, n_cols):
        ws.merge_cells(start_row=row, start_column=col,
                       end_row=row,   end_column=col + n_cols - 1)
        c = ws.cell(row=row, column=col, value=value)
        c.font = font(bold=True, color="FFFFFF", size=11)
        c.fill = fill(DARK); c.alignment = CENTRE
        return c

    # -------------------------------------------------------------------------
    # Helper: write a distribution table onto ws starting at (start_row, 1)
    # Returns next available row after the table.
    # -------------------------------------------------------------------------
    def write_dist_table(ws, df_table, section_title, start_row):
        cats   = list(df_table.index)
        months = [c for c in df_table.columns if c != "TOTAL"]
        n_cols = 1 + len(months) + 1   # label + months + TOTAL

        # Section title
        title_cell(ws, start_row, 1, section_title, n_cols)
        r = start_row + 1

        # Header
        hdr_cell(ws, r, 1, df_table.index.name or "Category")
        for c_off, m in enumerate(months):
            hdr_cell(ws, r, 2 + c_off, m)
        hdr_cell(ws, r, 2 + len(months), "TOTAL")
        r += 1

        # Data rows
        for row_off, cat in enumerate(cats):
            row_fill = fill(ALT) if row_off % 2 else fill(WHITE)
            c0 = ws.cell(row=r, column=1, value=cat)
            c0.font = font(bold=True); c0.fill = row_fill
            c0.alignment = LEFT; c0.border = BORDER

            for c_off, month in enumerate(months):
                cell = ws.cell(row=r, column=2 + c_off, value=int(df_table.loc[cat, month]))
                cell.font = font(); cell.fill = row_fill
                cell.alignment = CENTRE; cell.border = BORDER

            tot_col  = 2 + len(months)
            first_l  = get_column_letter(2)
            last_l   = get_column_letter(1 + len(months))
            tc = ws.cell(row=r, column=tot_col,
                         value=f"=SUM({first_l}{r}:{last_l}{r})")
            tc.font = font(bold=True); tc.fill = fill(LBLUE)
            tc.alignment = CENTRE; tc.border = BORDER
            r += 1

        # TOTAL row
        c0 = ws.cell(row=r, column=1, value="TOTAL")
        c0.font = font(bold=True); c0.fill = fill(LBLUE)
        c0.alignment = LEFT; c0.border = BORDER
        for c_off, month in enumerate(months):
            col   = 2 + c_off
            col_l = get_column_letter(col)
            data_start = start_row + 2
            cell = ws.cell(row=r, column=col,
                           value=f"=SUM({col_l}{data_start}:{col_l}{r - 1})")
            cell.font = font(bold=True); cell.fill = fill(LBLUE)
            cell.alignment = CENTRE; cell.border = BORDER

        tot_col = 2 + len(months)
        col_l   = get_column_letter(tot_col)
        data_start = start_row + 2
        gc = ws.cell(row=r, column=tot_col,
                     value=f"=SUM({col_l}{data_start}:{col_l}{r - 1})")
        gc.font = font(bold=True); gc.fill = fill(LBLUE)
        gc.alignment = CENTRE; gc.border = BORDER

        # Column widths
        ws.column_dimensions["A"].width = 18
        for c_off in range(len(months)):
            ws.column_dimensions[get_column_letter(2 + c_off)].width = 11
        ws.column_dimensions[get_column_letter(2 + len(months))].width = 9

        return r + 2   # blank row gap

    # -------------------------------------------------------------------------
    # One sheet per hospital: Locations
    # -------------------------------------------------------------------------
    def write_locations_sheet(ws, hosp_name, window_str, sub_period_info,
                               locs_pop, locs_infra):
        # Determine union of unique location names across both sheets
        pop_names   = set(locs_pop["location"])   if len(locs_pop)   else set()
        infra_names = set(locs_infra["location"]) if len(locs_infra) else set()
        all_names   = pop_names | infra_names

        # Build a combined reference DF (use pop for coords; infra for any not in pop)
        ref = pd.concat([locs_pop[["location","latitude","longitude"]],
                         locs_infra[["location","latitude","longitude"]]],
                        ignore_index=True).drop_duplicates("location")

        # Main title
        title_cell(ws, 1, 1,
                   f"{hosp_name}  |  Catchment Locations  |  {window_str}", 4)

        # Sub-period info block
        r = 2
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
        c = ws.cell(row=r, column=1, value="Voronoi Sub-periods & Catchment Areas")
        c.font = font(bold=True, color="FFFFFF"); c.fill = fill(STEEL)
        c.alignment = LEFT
        r += 1
        for sp in sub_period_info:
            ws.cell(row=r, column=1, value=sp["label"]).font = font(bold=True)
            ws.cell(row=r, column=2, value=f"{sp['start']} → {sp['end']}").font = font()
            ws.cell(row=r, column=3, value=f"{sp['area_km2']:.4f} km²").font = font()
            r += 1
        r += 1

        # Locations header
        for col_idx, header in enumerate(["#", "Location", "Latitude", "Longitude"], 1):
            hdr_cell(ws, r, col_idx, header)
        r += 1

        for i, row_data in enumerate(ref.itertuples(index=False), start=1):
            row_fill = fill(ALT) if i % 2 else fill(WHITE)
            ws.cell(row=r, column=1, value=i).fill              = row_fill
            ws.cell(row=r, column=1).font                       = font()
            ws.cell(row=r, column=1).alignment                  = CENTRE
            ws.cell(row=r, column=1).border                     = BORDER
            ws.cell(row=r, column=2, value=row_data.location).fill  = row_fill
            ws.cell(row=r, column=2).font                       = font()
            ws.cell(row=r, column=2).alignment                  = LEFT
            ws.cell(row=r, column=2).border                     = BORDER
            ws.cell(row=r, column=3, value=round(row_data.latitude, 6)).fill  = row_fill
            ws.cell(row=r, column=3).font                       = font()
            ws.cell(row=r, column=3).alignment                  = CENTRE
            ws.cell(row=r, column=3).border                     = BORDER
            ws.cell(row=r, column=4, value=round(row_data.longitude, 6)).fill = row_fill
            ws.cell(row=r, column=4).font                       = font()
            ws.cell(row=r, column=4).alignment                  = CENTRE
            ws.cell(row=r, column=4).border                     = BORDER
            r += 1

        # Total count row
        c0 = ws.cell(row=r, column=1, value="Total")
        c0.font = font(bold=True); c0.fill = fill(LBLUE); c0.border = BORDER
        c1 = ws.cell(row=r, column=2, value=len(ref))
        c1.font = font(bold=True); c1.fill = fill(LBLUE)
        c1.alignment = CENTRE; c1.border = BORDER

        ws.column_dimensions["A"].width = 5
        ws.column_dimensions["B"].width = 24
        ws.column_dimensions["C"].width = 14
        ws.column_dimensions["D"].width = 14
        ws.row_dimensions[1].height = 18
        ws.freeze_panes = "A5"

    # -------------------------------------------------------------------------
    # Write all hospitals
    # -------------------------------------------------------------------------
    for res in hospital_results:
        key        = res["key"]
        hosp_name  = res["name"]
        window_str = res["window_str"]
        sp_info    = res["sub_period_info"]
        locs_pop   = res["locations_pop"]
        locs_infra = res["locations_infra"]
        pop_table  = res["pop_table"]
        infra_table= res["infra_table"]

        # --- Locations sheet ---
        ws_loc = wb.create_sheet(title=f"{key}_Locations")
        write_locations_sheet(ws_loc, hosp_name, window_str, sp_info,
                               locs_pop, locs_infra)

        # --- Combined distribution sheet (pop + infra stacked) ---
        ws_dist = wb.create_sheet(title=f"{key}_Distributions")
        title_cell(ws_dist, 1, 1,
                   f"{hosp_name}  |  Distribution Tables  |  {window_str}", 10)
        ws_dist.row_dimensions[1].height = 18

        next_row = write_dist_table(
            ws_dist, pop_table,
            f"Population Density Distribution — {hosp_name}  ({window_str})",
            start_row=3,
        )
        write_dist_table(
            ws_dist, infra_table,
            f"Infrastructure Distribution — {hosp_name}  ({window_str})",
            start_row=next_row,
        )
        ws_dist.freeze_panes = "B4"

    wb.save(str(output_path))


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main() -> None:
    print("=" * 70)
    print("Catchment Distribution Analysis v2")
    print("=" * 70)

    print("\n[1] Loading Gaza boundary...")
    gaza_union = load_gaza_boundary(GAZA_GEOJSON)

    print("[2] Loading location data...")
    df_pop  = pd.read_excel(DENSITY_PATH, sheet_name="Population")
    df_infr = pd.read_excel(DENSITY_PATH, sheet_name="Infrastructure")
    df_pop  = df_pop.dropna(subset=["latitude", "longitude"]).reset_index(drop=True)
    df_infr = df_infr.dropna(subset=["latitude", "longitude"]).reset_index(drop=True)
    print(f"    Population locations  : {len(df_pop)}")
    print(f"    Infrastructure locations: {len(df_infr)}")

    geod             = Geod(ellps="WGS84")
    all_hosp_results = []

    for cfg in HOSPITAL_CONFIGS:
        hosp_key  = cfg["key"]
        hosp_name = cfg["name"]
        win_start, win_end = cfg["window"]
        sub_periods        = cfg["sub_periods"]
        window_str = f"{win_start.date()} to {win_end.date()}"

        print(f"\n{'='*70}")
        print(f"Hospital : {hosp_name}")
        print(f"Window   : {window_str}")

        # Compute one catchment per sub-period
        sub_catchments: Dict[int, Polygon] = {}
        sp_info: List[Dict] = []
        for idx, sp in enumerate(sub_periods):
            poly       = compute_catchment(hosp_name, sp["open"], gaza_union)
            area_m2, _ = geod.geometry_area_perimeter(poly)
            area_km2   = abs(area_m2) / 1e6
            sub_catchments[idx] = poly
            sp_info.append({
                "label"   : sp["label"],
                "start"   : sp["start"].date(),
                "end"     : sp["end"].date(),
                "area_km2": area_km2,
            })
            print(f"  Sub-period {idx+1} [{sp['label']}]: {area_km2:.4f} km²  "
                  f"({len(sp['open'])} hospitals in Voronoi)")

        # Month dispatcher
        window_months = months_in_window(win_start, win_end)
        if len(sub_periods) == 1:
            def make_single(sc):
                return lambda m: sc[0]
            cfn = make_single(sub_catchments)
        else:
            def make_multi(sp_list, sc, win):
                return lambda m: catchment_for_month(m, sp_list, sc, win)
            cfn = make_multi(sub_periods, sub_catchments, (win_start, win_end))

        # Union catchment for location list (use SP1 for single-period hospitals)
        # For EGH: show all locations appearing in ANY sub-period
        union_catchment = unary_union(list(sub_catchments.values()))

        locs_pop  = find_locations_in_catchment(df_pop,  union_catchment)
        locs_infra= find_locations_in_catchment(df_infr, union_catchment)

        # Print location lists
        print(f"\n  Locations in catchment (Population sheet) [{len(locs_pop)}]:")
        for loc in locs_pop["location"]:
            print(f"    • {loc}")
        print(f"\n  Locations in catchment (Infrastructure sheet) [{len(locs_infra)}]:")
        for loc in locs_infra["location"]:
            print(f"    • {loc}")

        # Build distribution tables
        pop_table = build_distribution_table(
            cfn, df_pop, window_months, POP_ORDER, "Pop Density Level"
        )
        infra_table = build_distribution_table(
            cfn, df_infr, window_months, INFRA_ORDER, "Infrastructure Level"
        )

        print(f"\n  Population Density Distribution:")
        print(pop_table.to_string())
        print(f"\n  Infrastructure Distribution:")
        print(infra_table.to_string())

        all_hosp_results.append({
            "key"           : hosp_key,
            "name"          : hosp_name,
            "window_str"    : window_str,
            "sub_period_info": sp_info,
            "locations_pop" : locs_pop[["location","latitude","longitude"]].reset_index(drop=True),
            "locations_infra": locs_infra[["location","latitude","longitude"]].reset_index(drop=True),
            "pop_table"     : pop_table,
            "infra_table"   : infra_table,
        })

    print(f"\n{'='*70}")
    print(f"[3] Writing Excel → {OUTPUT_PATH.name} ...")
    write_excel(all_hosp_results, OUTPUT_PATH)
    print("    Done.")
    print("=" * 70)


if __name__ == "__main__":
    main()
