from datetime import datetime, timedelta
from pathlib import Path
from typing import Dict, List, Tuple, Optional, Any

import numpy as np
import pandas as pd
import geopandas as gpd
from shapely.geometry import Point, Polygon, MultiPolygon, MultiPoint, box
from shapely.ops import unary_union
from pyproj import Geod
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Optional: shapely voronoi (faster when available)
try:
    from shapely.ops import voronoi_diagram
    from shapely import geometry as shapely_geom
    _HAS_SHAPELY_VORONOI = True
except ImportError:
    voronoi_diagram = None
    shapely_geom = None
    _HAS_SHAPELY_VORONOI = False

from scipy.spatial import Voronoi
import folium
from folium.plugins import HeatMap

# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------
BASE_PATH = Path(__file__).parent.resolve()
ACLED_PATH = BASE_PATH / "ACLED_May_09_25_Gaza.xlsx"
GAZA_GEOJSON = BASE_PATH / "pse_admin2.geojson"
OUTPUT_DIR = BASE_PATH

CRS_WGS84 = "EPSG:4326"
CRS_WEBMERC = "EPSG:3857"
CATCHMENT_DISTANCE_KM = 5.0
VORONOI_BUFFER_METERS = 15000
GAZA_AREA_KM2 = 364.871          # Official total area of Gaza Strip

# ---------------------------------------------------------------------------
# Hardcoded Hospital Availability Timeline
#
# Each entry: (hospital_name, lon, lat, [(date, status), ...])
# Coordinates are approximate centroids for Gaza hospitals.
# ---------------------------------------------------------------------------

# fmt: off
HOSPITAL_DEFINITIONS = [
    {
        "name": "Al Shifa Medical Hospital",
        "lon": 34.44274363,
        "lat": 31.52369093,
        "timeline": [
            (datetime(2023, 10,  1), "Open"),
            (datetime(2023, 11, 21), "Closed"),
            (datetime(2023, 12,  6), "Open"),
            (datetime(2024,  3, 18), "Closed"),
            (datetime(2024,  9,  1), "Open"),
        ],
    },
    {
        "name": "Al-Quds Hospital",
        "lon": 34.4297735,
        "lat": 31.5056094,
        "timeline": [
            (datetime(2023, 10,  1), "Open"),
            (datetime(2023, 11,  5), "Closed"),
            (datetime(2025,  3, 19), "Open"),
        ],
    },
    {
        "name": "Nasser Hospital",
        "lon": 34.29193795,
        "lat": 31.34689036,
        "timeline": [
            (datetime(2023, 10,  1), "Open"),
            (datetime(2024,  2, 20), "Closed"),
            (datetime(2024,  5,  3), "Open"),
        ],
    },
    {
        "name": "European Hospital",
        "lon": 34.31925517,
        "lat": 31.303174,
        "timeline": [
            (datetime(2023, 10,  1), "Open"),
            (datetime(2024,  7,  1), "Closed"),
            (datetime(2024,  8, 25), "Open"),
            (datetime(2025,  5, 13), "Closed"),
        ],
    },
    {
        "name": "Kuwait Hospital",
        "lon": 34.24915904,
        "lat": 31.28812189,
        "timeline": [
            (datetime(2023, 10,  1), "Open"),
        ],
    },
    {
        "name": "Al-Aqsa Hospital",
        "lon": 34.36,
        "lat": 31.419969,
        "timeline": [
            (datetime(2023, 10,  1), "Open"),
            (datetime(2024,  8,  1), "Closed"),
        ],
    },
]
# fmt: on


def build_hospitals_df() -> Tuple[pd.DataFrame, Dict[str, List[Tuple[datetime, str]]]]:
    """
    Build hospitals DataFrame and availability timeline from hardcoded definitions.

    Returns:
        hospitals_df          : DataFrame with columns Hospital, lon, lat
        availability_timeline : Dict mapping hospital_name -> [(datetime, status), ...]
    """
    rows = []
    timeline: Dict[str, List[Tuple[datetime, str]]] = {}

    for defn in HOSPITAL_DEFINITIONS:
        name = defn["name"]
        rows.append({"Hospital": name, "lon": defn["lon"], "lat": defn["lat"]})
        # Timeline is already sorted ascending and alternates Open/Closed
        timeline[name] = defn["timeline"]

    hospitals_df = pd.DataFrame(rows)
    return hospitals_df, timeline


# ---------------------------------------------------------------------------
# Availability Helpers
# ---------------------------------------------------------------------------

def get_hospital_status_at_date(
    availability_timeline: Dict[str, List[Tuple[datetime, str]]],
    hospital_name: str,
    date: datetime,
) -> str:
    """Return 'Open' or 'Closed' for a hospital on the given date."""
    if hospital_name not in availability_timeline:
        return "Closed"
    changes = availability_timeline[hospital_name]
    last_status = "Open"
    for dt, status in changes:
        if dt <= date:
            last_status = status
        else:
            break
    return last_status


def get_all_status_change_periods(
    hospitals_df: pd.DataFrame,
    availability_timeline: Dict[str, List[Tuple[datetime, str]]],
) -> List[Tuple[datetime, datetime]]:
    """
    Identify contiguous periods where the set of open hospitals does not change.
    Returns list of (period_start, period_end) tuples.
    """
    all_change_dates: set = set()
    for hosp_name in hospitals_df["Hospital"]:
        if hosp_name in availability_timeline:
            for dt, _ in availability_timeline[hosp_name]:
                all_change_dates.add(dt)

    sorted_dates = sorted(all_change_dates)
    if not sorted_dates:
        return []

    periods = []
    for i in range(len(sorted_dates) - 1):
        period_start = sorted_dates[i]
        period_end = sorted_dates[i + 1] - timedelta(days=1)
        periods.append((period_start, period_end))

    # Final period extends 1 year past the last change date
    final_end = sorted_dates[-1] + timedelta(days=365)
    periods.append((sorted_dates[-1], final_end))

    return periods


def get_open_hospitals_at_date(
    hospitals_df: pd.DataFrame,
    availability_timeline: Dict[str, List[Tuple[datetime, str]]],
    check_date: datetime,
) -> List[Tuple[str, float, float]]:
    """
    Return list of (hospital_name, lat, lon) for hospitals open on check_date.
    """
    open_hospitals = []
    for _, row in hospitals_df.iterrows():
        name = str(row["Hospital"]).strip()
        status = get_hospital_status_at_date(availability_timeline, name, check_date)
        if status == "Open":
            open_hospitals.append((name, float(row["lat"]), float(row["lon"])))
    return open_hospitals


# ---------------------------------------------------------------------------
# Gaza Boundary
# ---------------------------------------------------------------------------

def load_gaza_boundary(path: Path) -> Any:
    """
    Load Gaza outer boundary from GeoJSON.

    The file may contain multiple features (e.g. governorate polygons).
    We dissolve all features into a single union to obtain the outer border only.
    """
    gdf = gpd.read_file(path)
    if gdf.crs is None:
        gdf = gdf.set_crs(CRS_WGS84)
    gdf = gdf.to_crs(CRS_WGS84)
    # Dissolve all sub-regions into one outer boundary
    outer_boundary = gdf.unary_union
    return outer_boundary


# ---------------------------------------------------------------------------
# Catchment Area Method
# ---------------------------------------------------------------------------

def catchment_area_method(
    open_hospitals: List[Tuple[str, float, float]],
    gaza_union: Any,
    distance_cap_km: float = CATCHMENT_DISTANCE_KM,
) -> Dict[str, Tuple[Any, float]]:
    """
    Compute catchment areas: Voronoi region clipped to Gaza boundary,
    further capped to a 5 km geodesic buffer around each hospital.

    Returns: Dict mapping hospital_name -> (polygon_wgs84, area_km2)
    """
    if not open_hospitals:
        return {}

    hosp_gdf = gpd.GeoDataFrame(
        {"Hospital": [h[0] for h in open_hospitals]},
        geometry=[Point(h[2], h[1]) for h in open_hospitals],
        crs=CRS_WGS84,
    )
    hosp_proj = hosp_gdf.to_crs(CRS_WEBMERC)

    gaza_proj = gpd.GeoSeries([gaza_union], crs=CRS_WGS84).to_crs(CRS_WEBMERC)
    minx, miny, maxx, maxy = gaza_proj.total_bounds
    bbox_proj = box(
        minx - VORONOI_BUFFER_METERS,
        miny - VORONOI_BUFFER_METERS,
        maxx + VORONOI_BUFFER_METERS,
        maxy + VORONOI_BUFFER_METERS,
    )

    if _HAS_SHAPELY_VORONOI:
        try:
            multip = shapely_geom.MultiPoint(
                [(pt.x, pt.y) for pt in hosp_proj.geometry]
            )
            vor = voronoi_diagram(multip, envelope=bbox_proj, tolerance=0.0)
            polys_proj = _extract_voronoi_polygons(vor, bbox_proj, hosp_proj, hosp_gdf)
        except Exception:
            polys_proj = _scipy_voronoi_clipped(hosp_proj, hosp_gdf, bbox_proj, gaza_union)
    else:
        polys_proj = _scipy_voronoi_clipped(hosp_proj, hosp_gdf, bbox_proj, gaza_union)

    geod = Geod(ellps="WGS84")
    result: Dict[str, Tuple[Any, float]] = {}

    for hosp_name, poly in polys_proj.items():
        if poly is None or poly.is_empty:
            result[hosp_name] = (Polygon(), 0.0)
            continue

        poly_clipped = poly.intersection(gaza_union)
        if poly_clipped.is_empty:
            result[hosp_name] = (Polygon(), 0.0)
            continue

        # 5 km geodesic buffer cap
        hosp_row = hosp_gdf[hosp_gdf["Hospital"] == hosp_name]
        if hosp_row.empty:
            result[hosp_name] = (poly_clipped, 0.0)
            continue

        pt = hosp_row.geometry.iloc[0]
        angles = np.linspace(0, 360, 128)
        circle_pts = []
        for a in angles:
            lon2, lat2, _ = geod.fwd(pt.x, pt.y, a, distance_cap_km * 1000)
            circle_pts.append((lon2, lat2))
        circle_poly = Polygon(circle_pts)

        capped = poly_clipped.intersection(circle_poly)
        if capped.is_empty:
            result[hosp_name] = (Polygon(), 0.0)
            continue
        if isinstance(capped, MultiPolygon):
            capped = max(capped.geoms, key=lambda p: p.area)

        area_m2, _ = geod.geometry_area_perimeter(capped)
        area_km2 = abs(area_m2) / 1e6
        result[hosp_name] = (capped, area_km2)

    return result


def _extract_voronoi_polygons(vor, bbox_proj, hosp_proj, hosp_gdf) -> Dict[str, Any]:
    """Extract polygons from shapely voronoi_diagram and assign to hospitals."""
    poly_list = []
    if hasattr(vor, "geoms"):
        for g in vor.geoms:
            if isinstance(g, (Polygon, MultiPolygon)):
                poly_list.append(g)
    elif isinstance(vor, (Polygon, MultiPolygon)):
        poly_list = [vor]

    if not poly_list:
        return {}

    polys_proj = gpd.GeoDataFrame(geometry=poly_list, crs=CRS_WEBMERC)
    assigned: List[Tuple[str, Any]] = []
    hosp_pts_proj = hosp_proj.geometry

    for poly in polys_proj.geometry:
        if poly is None or poly.is_empty:
            continue
        rep = poly.representative_point()
        dists = hosp_pts_proj.distance(rep)
        nearest_idx = int(dists.idxmin())
        hosp_name = hosp_gdf.iloc[nearest_idx]["Hospital"]
        poly_wgs = gpd.GeoSeries([poly], crs=CRS_WEBMERC).to_crs(CRS_WGS84).iloc[0]
        assigned.append((hosp_name, poly_wgs))

    result: Dict[str, Any] = {}
    for hosp in hosp_gdf["Hospital"].values:
        polys_for = [p for (h, p) in assigned if h == hosp]
        result[hosp] = unary_union(polys_for) if polys_for else Polygon()
    return result


def _scipy_voronoi_clipped(hosp_proj, hosp_gdf, bbox_proj, gaza_union) -> Dict[str, Any]:
    """Fallback: SciPy Voronoi with bounding box clipping."""
    coords = np.array([(pt.x, pt.y) for pt in hosp_proj.geometry])
    vor = Voronoi(coords)
    hosp_names = hosp_gdf["Hospital"].values

    def make_bounded_poly(site: np.ndarray) -> Polygon:
        bbox_coords = np.array(bbox_proj.exterior.coords)
        pts = np.vstack([site, bbox_coords])
        multipoint = MultiPoint([tuple(p) for p in pts])
        return multipoint.convex_hull.intersection(bbox_proj)

    proj_polys = []
    for pt_idx, region_idx in enumerate(vor.point_region):
        region = vor.regions[region_idx]
        if not region or -1 in region:
            poly = make_bounded_poly(coords[pt_idx])
        else:
            try:
                verts = [vor.vertices[i] for i in region]
                poly = Polygon(verts).intersection(bbox_proj)
            except Exception:
                poly = make_bounded_poly(coords[pt_idx])
        proj_polys.append(poly)

    polys_wgs = (
        gpd.GeoDataFrame(geometry=proj_polys, crs=CRS_WEBMERC).to_crs(CRS_WGS84)
    )
    result: Dict[str, Any] = {}
    for i, hosp_name in enumerate(hosp_names):
        try:
            clipped = polys_wgs.iloc[i].geometry.intersection(gaza_union)
        except Exception:
            clipped = Polygon()
        result[hosp_name] = clipped if clipped and not clipped.is_empty else Polygon()

    # Assign leftover area to nearest hospital
    covered = unary_union([g for g in result.values() if g and not g.is_empty])
    leftover = gaza_union.difference(covered) if covered else gaza_union
    if leftover and not leftover.is_empty:
        pieces = [leftover] if isinstance(leftover, Polygon) else list(leftover.geoms)
        hosp_pts_proj = hosp_proj.geometry
        for piece in pieces:
            if piece.is_empty:
                continue
            piece_proj = (
                gpd.GeoSeries([piece], crs=CRS_WGS84).to_crs(CRS_WEBMERC).iloc[0]
            )
            rep = piece_proj.representative_point()
            dists = hosp_pts_proj.distance(rep)
            nearest_idx = int(dists.idxmin())
            n = hosp_names[nearest_idx]
            cur = result.get(n, Polygon())
            result[n] = cur.union(piece) if cur and not cur.is_empty else piece

    return result


# ---------------------------------------------------------------------------
# ACLED Helpers
# ---------------------------------------------------------------------------

def load_acled(path: Path) -> pd.DataFrame:
    """Load ACLED attack data, auto-detecting date/lat/lon columns."""
    df = pd.read_excel(path)
    date_col = lat_col = lon_col = None
    for c in df.columns:
        cl = str(c).lower()
        if not date_col and ("event_date" in cl or "date" in cl or "eventdate" in cl):
            date_col = c
        if not lat_col and ("lat" in cl or "latitude" in cl or cl == "y"):
            lat_col = c
        if not lon_col and ("lon" in cl or "longitude" in cl or cl == "x"):
            lon_col = c

    if not (date_col and lat_col and lon_col):
        raise ValueError(
            f"Cannot find date/lat/lon columns in ACLED. Columns found: {list(df.columns)}"
        )

    df = df.rename(columns={date_col: "_date", lat_col: "_lat", lon_col: "_lon"})
    df["_date"] = pd.to_datetime(df["_date"]).dt.date
    df["_lat"] = pd.to_numeric(df["_lat"], errors="coerce")
    df["_lon"] = pd.to_numeric(df["_lon"], errors="coerce")
    return df.dropna(subset=["_date", "_lat", "_lon"]).copy()


def pts_to_heatlist(
    acled_df: pd.DataFrame,
    period_start: datetime,
    period_end: datetime,
    round_decimals: int = 5,
) -> Tuple[List[List[float]], float]:
    """
    Convert ACLED rows in [period_start, period_end] to HeatMap input.
    Co-located events are aggregated so their weight increases.

    Returns (heat_list, max_weight).
    """
    mask = (acled_df["_date"] >= period_start.date()) & (
        acled_df["_date"] <= period_end.date()
    )
    subset = acled_df[mask].copy()
    if subset.empty:
        return [], 1.0

    subset["lon_r"] = subset["_lon"].round(round_decimals)
    subset["lat_r"] = subset["_lat"].round(round_decimals)
    grouped = subset.groupby(["lat_r", "lon_r"]).size().reset_index(name="count")
    heat = grouped.apply(
        lambda r: [float(r["lat_r"]), float(r["lon_r"]), float(r["count"])], axis=1
    ).tolist()
    max_weight = float(grouped["count"].max()) if not grouped.empty else 1.0
    return heat, max_weight


def attack_count_method(
    acled_df: pd.DataFrame,
    polygon: Any,
    period_start: datetime,
    period_end: datetime,
) -> int:
    """Count ACLED events inside polygon during the given period."""
    mask = (acled_df["_date"] >= period_start.date()) & (
        acled_df["_date"] <= period_end.date()
    )
    subset = acled_df[mask]
    count = 0
    for _, row in subset.iterrows():
        pt = Point(row["_lon"], row["_lat"])
        if pt.within(polygon):
            count += 1
    return count


# ---------------------------------------------------------------------------
# HTML Map Output
# ---------------------------------------------------------------------------

def output_html_map(
    open_hospitals: List[Tuple[str, float, float]],
    catchments: Dict[str, Tuple[Any, float]],
    attack_counts: Dict[str, int],
    acled_df: pd.DataFrame,
    period_start: datetime,
    period_end: datetime,
    output_path: Path,
) -> None:
    """
    Write an HTML folium map for one time period showing:
    - Hospital markers
    - Catchment polygons (coloured per hospital, tooltip shows area + attack count)
    - Heatmap layer of ACLED attacks
    - Info box summarising the period
    """
    if not open_hospitals:
        return

    center_lat = sum(h[1] for h in open_hospitals) / len(open_hospitals)
    center_lon = sum(h[2] for h in open_hospitals) / len(open_hospitals)
    m = folium.Map(location=[center_lat, center_lon], zoom_start=11, tiles="OpenStreetMap")

    # Heatmap layer
    heat_list, max_weight = pts_to_heatlist(acled_df, period_start, period_end)
    if heat_list:
        HeatMap(
            heat_list,
            max_val=max_weight,
            radius=15,
            blur=20,
            max_zoom=13,
            min_opacity=0.3,
            gradient={0.4: "blue", 0.6: "lime", 0.7: "yellow", 0.8: "orange", 1.0: "red"},
        ).add_to(m)

    colors = ["#e41a1c", "#377eb8", "#4daf4a", "#984ea3", "#ff7f00"]

    for i, (hosp_name, lat, lon) in enumerate(open_hospitals):
        color = colors[i % len(colors)]
        folium.CircleMarker(
            location=[lat, lon],
            radius=10,
            color=color,
            fill=True,
            fillColor=color,
            fillOpacity=0.9,
            popup=f"<b>{hosp_name}</b>",
            weight=2,
        ).add_to(m)

    for i, (hosp_name, (poly, area_km2)) in enumerate(catchments.items()):
        if poly is None or poly.is_empty:
            continue
        color = colors[i % len(colors)]
        attacks = attack_counts.get(hosp_name, 0)
        tooltip_text = f"{hosp_name}<br>Area: {area_km2:.2f} km²<br>Attacks: {attacks}"

        if isinstance(poly, MultiPolygon):
            for geom in poly.geoms:
                _add_polygon_to_map(m, geom, color, tooltip_text)
        else:
            _add_polygon_to_map(m, poly, color, tooltip_text)

    total_attacks = sum(attack_counts.values())
    title_html = f"""
    <div style="position: fixed; top: 10px; left: 50px; width: 400px; padding: 10px;
                background: white; border: 2px solid grey; z-index: 9999; font-size: 14px;">
    <b>Hospital Catchment Areas</b><br>
    Period: {period_start.strftime('%Y-%m-%d')} to {period_end.strftime('%Y-%m-%d')}<br>
    Open Hospitals: {len(open_hospitals)}<br>
    Total Attacks in Period: {total_attacks}
    </div>"""
    m.get_root().html.add_child(folium.Element(title_html))
    m.save(str(output_path))


def _add_polygon_to_map(m: folium.Map, poly: Polygon, color: str, tooltip: str) -> None:
    """Add a single Polygon to a folium Map."""
    if poly.is_empty or not poly.exterior:
        return
    coords = [[c[1], c[0]] for c in poly.exterior.coords]
    folium.Polygon(
        locations=coords,
        color=color,
        fill=True,
        fillColor=color,
        fillOpacity=0.3,
        weight=2,
        tooltip=tooltip,
    ).add_to(m)


# ---------------------------------------------------------------------------
# Excel Summary Output
# ---------------------------------------------------------------------------

def output_excel_summary(
    all_period_data: List[Dict],
    output_path: Path,
) -> None:
    """
    Write a single Excel workbook summarising all periods.

    Sheet layout
    ------------
    One sheet per time period, named by its date range.
    Each sheet has one row per hospital that was open during that period, plus
    a TOTAL row at the bottom.

    Columns
    -------
    Hospital | Status | Catchment Area (km²) | % of Gaza | Attacks in Catchment | Total Attacks in Period
    """
    wb = Workbook()
    wb.remove(wb.active)          # remove default blank sheet

    # --- Style helpers ---
    HEADER_FILL   = PatternFill("solid", start_color="1F4E79", end_color="1F4E79")
    TOTAL_FILL    = PatternFill("solid", start_color="D6E4F0", end_color="D6E4F0")
    CLOSED_FILL   = PatternFill("solid", start_color="F2F2F2", end_color="F2F2F2")
    WHITE_FILL    = PatternFill("solid", start_color="FFFFFF", end_color="FFFFFF")

    HDR_FONT      = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    BODY_FONT     = Font(name="Arial", size=10)
    TOTAL_FONT    = Font(name="Arial", bold=True, size=10)
    CLOSED_FONT   = Font(name="Arial", size=10, color="999999", italic=True)

    CENTRE        = Alignment(horizontal="center", vertical="center")
    LEFT          = Alignment(horizontal="left",   vertical="center")
    RIGHT         = Alignment(horizontal="right",  vertical="center")

    thin          = Side(style="thin", color="BFBFBF")
    BORDER        = Border(left=thin, right=thin, top=thin, bottom=thin)

    COL_HEADERS = [
        "Hospital",
        "Status",
        "Catchment Area (km²)",
        "% of Gaza",
        "Attacks in Catchment",
        "Total Attacks in Period",
    ]
    COL_WIDTHS = [28, 10, 22, 12, 22, 24]

    all_hospital_names = [d["name"] for d in HOSPITAL_DEFINITIONS]

    for period_data in all_period_data:
        period_start : datetime = period_data["period_start"]
        period_end   : datetime = period_data["period_end"]
        catchments   : Dict[str, Tuple[Any, float]] = period_data["catchments"]
        attack_counts: Dict[str, int]               = period_data["attack_counts"]
        open_set     : set                          = period_data["open_set"]
        total_attacks: int                          = period_data["total_attacks"]

        sheet_name = (
            f"{period_start.strftime('%Y%m%d')}-{period_end.strftime('%Y%m%d')}"
        )
        ws = wb.create_sheet(title=sheet_name)
        ws.freeze_panes = "A2"

        # ---- Period header (merged across all columns) ----
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(COL_HEADERS))
        title_cell = ws.cell(row=1, column=1,
            value=(
                f"Period: {period_start.strftime('%Y-%m-%d')} "
                f"to {period_end.strftime('%Y-%m-%d')}"
            )
        )
        title_cell.font      = Font(name="Arial", bold=True, size=11, color="FFFFFF")
        title_cell.fill      = PatternFill("solid", start_color="0D2B55", end_color="0D2B55")
        title_cell.alignment = CENTRE

        # ---- Column headers ----
        for col_idx, header in enumerate(COL_HEADERS, start=1):
            cell            = ws.cell(row=2, column=col_idx, value=header)
            cell.font       = HDR_FONT
            cell.fill       = HEADER_FILL
            cell.alignment  = CENTRE
            cell.border     = BORDER

        # ---- Data rows — one per hospital (open and closed) ----
        data_start_row = 3
        for row_offset, hosp_name in enumerate(all_hospital_names):
            r        = data_start_row + row_offset
            is_open  = hosp_name in open_set
            status   = "Open" if is_open else "Closed"

            if is_open:
                _, area_km2  = catchments.get(hosp_name, (None, 0.0))
                hosp_attacks = attack_counts.get(hosp_name, 0)
                fill         = WHITE_FILL
                font         = BODY_FONT
            else:
                area_km2     = 0.0
                hosp_attacks = 0
                fill         = CLOSED_FILL
                font         = CLOSED_FONT

            area_col    = get_column_letter(3)   # C
            pct_formula = f"={area_col}{r}/{GAZA_AREA_KM2}"

            values = [
                hosp_name,
                status,
                round(area_km2, 4) if is_open else "—",
                pct_formula       if is_open else "—",
                hosp_attacks      if is_open else "—",
                total_attacks,
            ]
            aligns = [LEFT, CENTRE, RIGHT, RIGHT, RIGHT, RIGHT]

            for col_idx, (val, aln) in enumerate(zip(values, aligns), start=1):
                cell            = ws.cell(row=r, column=col_idx, value=val)
                cell.font       = font
                cell.fill       = fill
                cell.alignment  = aln
                cell.border     = BORDER
                # Number formats
                if col_idx == 3 and is_open:
                    cell.number_format = "0.00"
                if col_idx == 4 and is_open:
                    cell.number_format = "0.0%"

        # ---- TOTAL row ----
        total_row    = data_start_row + len(all_hospital_names)
        open_rows    = [
            data_start_row + i
            for i, h in enumerate(all_hospital_names)
            if h in open_set
        ]
        area_sum_ref = (
            "+".join(f"C{r}" for r in open_rows) if open_rows else "0"
        )
        atk_sum_ref  = (
            "+".join(f"E{r}" for r in open_rows) if open_rows else "0"
        )

        total_values = [
            "TOTAL",
            f"{len(open_set)} open",
            f"={area_sum_ref}",
            f"=C{total_row}/{GAZA_AREA_KM2}",
            f"={atk_sum_ref}",
            total_attacks,
        ]
        total_aligns = [LEFT, CENTRE, RIGHT, RIGHT, RIGHT, RIGHT]

        for col_idx, (val, aln) in enumerate(zip(total_values, total_aligns), start=1):
            cell            = ws.cell(row=total_row, column=col_idx, value=val)
            cell.font       = TOTAL_FONT
            cell.fill       = TOTAL_FILL
            cell.alignment  = aln
            cell.border     = BORDER
            if col_idx == 3:
                cell.number_format = "0.00"
            if col_idx == 4:
                cell.number_format = "0.0%"

        # ---- Column widths ----
        for col_idx, width in enumerate(COL_WIDTHS, start=1):
            ws.column_dimensions[get_column_letter(col_idx)].width = width
        ws.row_dimensions[1].height = 18
        ws.row_dimensions[2].height = 16

    # ---- Index sheet (first sheet, lists all periods) ----
    index_ws = wb.create_sheet(title="Index", index=0)
    idx_hdrs = ["Period", "Start Date", "End Date", "Open Hospitals", "Total Attacks"]
    for col_idx, h in enumerate(idx_hdrs, start=1):
        cell           = index_ws.cell(row=1, column=col_idx, value=h)
        cell.font      = HDR_FONT
        cell.fill      = HEADER_FILL
        cell.alignment = CENTRE
        cell.border    = BORDER

    for r_offset, period_data in enumerate(all_period_data, start=2):
        ps   = period_data["period_start"]
        pe   = period_data["period_end"]
        vals = [
            f"{ps.strftime('%Y-%m-%d')} → {pe.strftime('%Y-%m-%d')}",
            ps.strftime("%Y-%m-%d"),
            pe.strftime("%Y-%m-%d"),
            len(period_data["open_set"]),
            period_data["total_attacks"],
        ]
        for col_idx, val in enumerate(vals, start=1):
            cell           = index_ws.cell(row=r_offset, column=col_idx, value=val)
            cell.font      = BODY_FONT
            cell.alignment = CENTRE if col_idx > 1 else LEFT
            cell.border    = BORDER

    for col_idx, width in enumerate([38, 14, 14, 16, 16], start=1):
        index_ws.column_dimensions[get_column_letter(col_idx)].width = width
    index_ws.freeze_panes = "A2"

    wb.save(str(output_path))


# ---------------------------------------------------------------------------
# Main Pipeline
# ---------------------------------------------------------------------------

def main() -> None:
    print("=" * 70)
    print("Hospital Catchment Area HTML Map Generation")
    print("=" * 70)

    # 1. Build hospital list and availability timeline from hardcoded data
    print("\n[1] Building hospital availability timeline from hardcoded data...")
    hospitals_df, availability_timeline = build_hospitals_df()
    print(f"    {len(hospitals_df)} hospitals defined")
    for _, row in hospitals_df.iterrows():
        events = availability_timeline[row['Hospital']]
        print(f"    • {row['Hospital']}  ({len(events)} timeline entries)")

    # 2. Load Gaza boundary (dissolve all sub-regions into outer border)
    print("\n[2] Loading Gaza boundary from GeoJSON (dissolving sub-regions)...")
    gaza_union = load_gaza_boundary(GAZA_GEOJSON)
    print("    Gaza boundary loaded and dissolved to outer border")

    # 3. Load ACLED
    print("\n[3] Loading ACLED attack data...")
    acled_df = load_acled(ACLED_PATH)
    print(f"    {len(acled_df)} attack events loaded")

    # 4. Identify status-change periods
    print("\n[4] Identifying status-change periods...")
    periods = get_all_status_change_periods(hospitals_df, availability_timeline)
    print(f"    {len(periods)} distinct periods found")

    # 5. Generate one HTML map per period + collect data for Excel
    print("\n[5] Generating HTML maps...")
    map_count     = 0
    all_period_data: List[Dict] = []

    for period_start, period_end in periods:
        open_hospitals = get_open_hospitals_at_date(
            hospitals_df, availability_timeline, period_start
        )

        if not open_hospitals:
            print(
                f"    Skipping {period_start.strftime('%Y-%m-%d')} → "
                f"{period_end.strftime('%Y-%m-%d')}: no open hospitals"
            )
            continue

        catchments = catchment_area_method(open_hospitals, gaza_union)

        attack_counts: Dict[str, int] = {}
        for hosp_name, (poly, _) in catchments.items():
            if poly and not poly.is_empty:
                attack_counts[hosp_name] = attack_count_method(
                    acled_df, poly, period_start, period_end
                )
            else:
                attack_counts[hosp_name] = 0

        total_attacks = sum(attack_counts.values())

        start_str  = period_start.strftime("%Y%m%d")
        end_str    = period_end.strftime("%Y%m%d")
        html_path  = OUTPUT_DIR / f"{start_str}_{end_str}_catchmaps.html"

        output_html_map(
            open_hospitals,
            catchments,
            attack_counts,
            acled_df,
            period_start,
            period_end,
            html_path,
        )

        # Accumulate data for Excel summary
        all_period_data.append({
            "period_start" : period_start,
            "period_end"   : period_end,
            "catchments"   : catchments,
            "attack_counts": attack_counts,
            "open_set"     : {h[0] for h in open_hospitals},
            "total_attacks": total_attacks,
        })

        map_count += 1
        open_names = ", ".join(h[0] for h in open_hospitals)
        print(
            f"    [{map_count}] {html_path.name}\n"
            f"         Open: {open_names}\n"
            f"         Total attacks in period: {total_attacks}"
        )

    # 6. Write Excel summary
    print("\n[6] Writing Excel summary...")
    excel_path = OUTPUT_DIR / "catchment_summary.xlsx"
    output_excel_summary(all_period_data, excel_path)
    print(f"    Saved: {excel_path.name}")

    print(f"\n[7] Done — {map_count} HTML maps + 1 Excel summary written to {OUTPUT_DIR}")
    print("=" * 70)


if __name__ == "__main__":
    main()
