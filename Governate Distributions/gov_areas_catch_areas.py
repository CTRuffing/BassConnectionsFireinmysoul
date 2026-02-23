"""
generate_catchment_excel.py

Reads all *_catchmaps.html files produced by the catchment area pipeline,
re-derives catchment polygons from the hospital open/close data, intersects
them with Gaza Strip governorates from pse_admin2.geojson, and writes a
single Excel spreadsheet summarising each timeline period.

Expected columns per row
------------------------
Timeline Range | Hospital | Total Catchment Area (km²) |
  <Gov1> Catchment (km²) | <Gov2> Catchment (km²) | ... |
  <Gov1> Total Area (km²) | <Gov2> Total Area (km²) | ...

Usage
-----
Edit the four path constants below, then run:
    python generate_catchment_excel.py

Dependencies
------------
geopandas, shapely, pyproj, openpyxl, pandas, scipy, numpy
(all already required by the original pipeline)
"""

from datetime import datetime, timedelta
from pathlib import Path
from typing import Dict, List, Tuple, Optional, Any
import re

import numpy as np
import pandas as pd
import geopandas as gpd
from shapely.geometry import Point, Polygon, MultiPolygon, box
from shapely.ops import unary_union
from pyproj import Geod
from scipy.spatial import Voronoi

try:
    from shapely.ops import voronoi_diagram
    from shapely import geometry as shapely_geom
    _HAS_SHAPELY_VORONOI = True
except ImportError:
    voronoi_diagram = None
    shapely_geom = None
    _HAS_SHAPELY_VORONOI = False

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# *** EDIT THESE PATHS ***
# ---------------------------------------------------------------------------
# Hospital availability Excel and Gaza boundary GeoJSON live in the parent folder
HOSP_PATH       = Path("..") / "Hospitals_OpenCloseoverTime.xlsx"
GAZA_GEOJSON    = Path("..") / "gaza_boundary.geojson"       # boundary used for clipping

# Governorate polygons GeoJSON lives in this folder
ADMIN2_GEOJSON  = Path("pse_admin2.geojson")

# Folder (in the parent directory) containing the *_catchmaps.html files
HTML_DIR        = Path("..") / "GeographicCRS_EPSG4326 -LatLong"

# Excel output will be written into the current (Governate Distributions) folder
OUTPUT_XLSX     = Path("catchment_summary.xlsx")

# Only governorates belonging to Gaza Strip will be used.
GAZA_ADM1_NAME  = "Gaza Strip"

# Must match the pipeline constants
CRS_WGS84           = "EPSG:4326"
CRS_WEBMERC         = "EPSG:3857"
CATCHMENT_DISTANCE_KM = 5.0
VORONOI_BUFFER_METERS = 15_000

# ---------------------------------------------------------------------------
# Helpers copied / adapted from the original pipeline (no imports from it)
# ---------------------------------------------------------------------------

def _parse_date(x: Any) -> Optional[datetime]:
    if x is None:
        return None
    try:
        if pd.isna(x):
            return None
    except (TypeError, ValueError):
        pass
    if isinstance(x, datetime):
        return x
    try:
        return pd.to_datetime(x).to_pydatetime()
    except Exception:
        return None


def read_hospitals_open_close(path: Path) -> Tuple[pd.DataFrame, Dict[str, List[Tuple[datetime, str]]]]:
    raw = pd.read_excel(path, header=None)
    hosp_col, lon_col, lat_col = 0, 1, 2
    schedule_cols = list(range(3, raw.shape[1]))

    dates_in_header = False
    if len(raw) >= 2 and schedule_cols:
        if _parse_date(raw.iloc[0, 3]) is not None:
            dates_in_header = True

    schedule_events: List[Tuple[datetime, str]] = []
    if dates_in_header:
        for i, col_idx in enumerate(schedule_cols):
            if col_idx >= raw.shape[1]:
                break
            status = "Open" if i % 2 == 0 else "Closed"
            dt = _parse_date(raw.iloc[1, col_idx])
            if dt is not None:
                schedule_events.append((dt, status))
        schedule_events = sorted(schedule_events, key=lambda x: x[0])
        compressed = []
        for dt, typ in schedule_events:
            if not compressed or compressed[-1][1] != typ:
                compressed.append((dt, typ))
        if not compressed:
            compressed = [(datetime(1900, 1, 1), "Open")]
        schedule_events = compressed
        data_start_row = 2
    else:
        data_start_row = 1

    hospitals_rows: List[Dict] = []
    availability_timeline: Dict[str, List[Tuple[datetime, str]]] = {}

    for raw_row_idx in range(data_start_row, len(raw)):
        name_val = raw.iloc[raw_row_idx, hosp_col]
        lon_val  = pd.to_numeric(raw.iloc[raw_row_idx, lon_col], errors="coerce")
        lat_val  = pd.to_numeric(raw.iloc[raw_row_idx, lat_col], errors="coerce")
        if pd.isna(lon_val) or pd.isna(lat_val):
            continue
        name = str(name_val).strip()
        if not name or name == "nan":
            continue
        hospitals_rows.append({"Hospital": name, "lon": float(lon_val), "lat": float(lat_val)})

        if dates_in_header:
            availability_timeline[name] = schedule_events
        else:
            events: List[Tuple[datetime, str]] = []
            for i, col_idx in enumerate(schedule_cols):
                if col_idx >= raw.shape[1]:
                    break
                status = "Open" if i % 2 == 0 else "Closed"
                val = raw.iloc[raw_row_idx, col_idx]
                dt = _parse_date(val)
                if dt is not None:
                    events.append((dt, status))
            events = sorted(events, key=lambda x: x[0])
            compressed_list: List[Tuple[datetime, str]] = []
            for dt, typ in events:
                if not compressed_list or compressed_list[-1][1] != typ:
                    compressed_list.append((dt, typ))
            if not compressed_list:
                compressed_list = [(datetime(1900, 1, 1), "Open")]
            availability_timeline[name] = compressed_list

    return pd.DataFrame(hospitals_rows), availability_timeline


def get_hospital_status_at_date(timeline, name, date):
    if name not in timeline:
        return "Closed"
    last = "Open"
    for dt, typ in timeline[name]:
        if dt <= date:
            last = typ
        else:
            break
    return last


def get_all_status_change_periods(hospitals_df, timeline):
    all_dates = set()
    for name in hospitals_df["Hospital"]:
        if name in timeline:
            for dt, _ in timeline[name]:
                all_dates.add(dt)
    sorted_dates = sorted(all_dates)
    if not sorted_dates:
        return []
    periods = []
    for i in range(len(sorted_dates) - 1):
        periods.append((sorted_dates[i], sorted_dates[i + 1] - timedelta(days=1)))
    periods.append((sorted_dates[-1], sorted_dates[-1] + timedelta(days=365)))
    return periods


def get_open_hospitals_at_date(hospitals_df, timeline, check_date):
    open_hospitals = []
    for _, row in hospitals_df.iterrows():
        name = str(row["Hospital"]).strip()
        if get_hospital_status_at_date(timeline, name, check_date) == "Open":
            open_hospitals.append((name, float(row["lat"]), float(row["lon"])))
    return open_hospitals


# ---------------------------------------------------------------------------
# Voronoi / catchment (identical logic to original pipeline)
# ---------------------------------------------------------------------------

def _extract_voronoi_polygons(vor, bbox_proj, hosp_proj, hosp_gdf):
    poly_list = []
    if hasattr(vor, "geoms"):
        for g in vor.geoms:
            if isinstance(g, (Polygon, MultiPolygon)):
                poly_list.append(g)
    elif isinstance(vor, (Polygon, MultiPolygon)):
        poly_list = [vor]
    if not poly_list:
        return {}
    polys_gdf = gpd.GeoDataFrame(geometry=poly_list, crs=CRS_WEBMERC)
    assigned = []
    hosp_pts_proj = hosp_proj.geometry
    for poly in polys_gdf.geometry:
        if poly is None or poly.is_empty:
            continue
        rep = poly.representative_point()
        dists = hosp_pts_proj.distance(rep)
        nearest_idx = int(dists.idxmin())
        hosp_name = hosp_gdf.iloc[nearest_idx]["Hospital"]
        poly_wgs = gpd.GeoSeries([poly], crs=CRS_WEBMERC).to_crs(CRS_WGS84).iloc[0]
        assigned.append((hosp_name, poly_wgs))
    result = {}
    for hosp in hosp_gdf["Hospital"].values:
        polys_for = [p for (h, p) in assigned if h == hosp]
        result[hosp] = unary_union(polys_for) if polys_for else Polygon()
    return result


def _scipy_voronoi_clipped(hosp_proj, hosp_gdf, bbox_proj, gaza_union):
    from shapely.geometry import MultiPoint
    coords = np.array([(pt.x, pt.y) for pt in hosp_proj.geometry])
    vor = Voronoi(coords)
    hosp_names = hosp_gdf["Hospital"].values

    def make_bounded_poly(site):
        bbox_coords = np.array(bbox_proj.exterior.coords)
        pts = np.vstack([site, bbox_coords])
        mp = MultiPoint([tuple(p) for p in pts])
        return mp.convex_hull.intersection(bbox_proj)

    proj_polys = []
    for pt_idx, region_idx in enumerate(vor.point_region):
        region = vor.regions[region_idx]
        if not region or -1 in region:
            poly = make_bounded_poly(coords[pt_idx])
        else:
            try:
                verts = [vor.vertices[i] for i in region]
                poly = Polygon(verts).intersection(bbox_proj)
            except Exception:
                poly = make_bounded_poly(coords[pt_idx])
        proj_polys.append(poly)

    polys_wgs = gpd.GeoDataFrame(geometry=proj_polys, crs=CRS_WEBMERC).to_crs(CRS_WGS84)
    result = {}
    for i, hosp_name in enumerate(hosp_names):
        try:
            clipped = polys_wgs.iloc[i].geometry.intersection(gaza_union)
        except Exception:
            clipped = Polygon()
        result[hosp_name] = clipped if clipped and not clipped.is_empty else Polygon()

    covered  = unary_union([g for g in result.values() if g and not g.is_empty])
    leftover = gaza_union.difference(covered) if covered else gaza_union
    if leftover and not leftover.is_empty:
        pieces = [leftover] if isinstance(leftover, Polygon) else list(leftover.geoms)
        hosp_pts_proj = hosp_proj.geometry
        for piece in pieces:
            if piece.is_empty:
                continue
            piece_proj = gpd.GeoSeries([piece], crs=CRS_WGS84).to_crs(CRS_WEBMERC).iloc[0]
            rep = piece_proj.representative_point()
            dists = hosp_pts_proj.distance(rep)
            nearest_idx = int(dists.idxmin())
            n = hosp_names[nearest_idx]
            cur = result.get(n, Polygon())
            result[n] = cur.union(piece) if cur and not cur.is_empty else piece
    return result


def catchment_area_method(open_hospitals, gaza_union, distance_cap_km=CATCHMENT_DISTANCE_KM):
    if not open_hospitals:
        return {}

    hosp_gdf = gpd.GeoDataFrame(
        {"Hospital": [h[0] for h in open_hospitals]},
        geometry=[Point(h[2], h[1]) for h in open_hospitals],
        crs=CRS_WGS84
    )
    hosp_proj = hosp_gdf.to_crs(CRS_WEBMERC)
    gaza_proj = gpd.GeoSeries([gaza_union], crs=CRS_WGS84).to_crs(CRS_WEBMERC)
    minx, miny, maxx, maxy = gaza_proj.total_bounds
    bbox_proj = box(
        minx - VORONOI_BUFFER_METERS, miny - VORONOI_BUFFER_METERS,
        maxx + VORONOI_BUFFER_METERS, maxy + VORONOI_BUFFER_METERS
    )

    if _HAS_SHAPELY_VORONOI:
        try:
            multip = shapely_geom.MultiPoint([(pt.x, pt.y) for pt in hosp_proj.geometry])
            vor = voronoi_diagram(multip, envelope=bbox_proj, tolerance=0.0)
            polys_proj = _extract_voronoi_polygons(vor, bbox_proj, hosp_proj, hosp_gdf)
        except Exception:
            polys_proj = _scipy_voronoi_clipped(hosp_proj, hosp_gdf, bbox_proj, gaza_union)
    else:
        polys_proj = _scipy_voronoi_clipped(hosp_proj, hosp_gdf, bbox_proj, gaza_union)

    geod = Geod(ellps="WGS84")
    result: Dict[str, Tuple[Any, float]] = {}

    for hosp_name, poly in polys_proj.items():
        if poly is None or poly.is_empty:
            result[hosp_name] = (Polygon(), 0.0)
            continue
        poly_clipped = poly.intersection(gaza_union)
        if poly_clipped.is_empty:
            result[hosp_name] = (Polygon(), 0.0)
            continue
        hosp_row = hosp_gdf[hosp_gdf["Hospital"] == hosp_name]
        if hosp_row.empty:
            result[hosp_name] = (poly_clipped, 0.0)
            continue
        pt = hosp_row.geometry.iloc[0]
        angles = np.linspace(0, 360, 128)
        circle_pts = []
        for a in angles:
            lon2, lat2, _ = geod.fwd(pt.x, pt.y, a, distance_cap_km * 1000)
            circle_pts.append((lon2, lat2))
        circle_poly = Polygon(circle_pts)
        capped = poly_clipped.intersection(circle_poly)
        if capped.is_empty:
            result[hosp_name] = (Polygon(), 0.0)
            continue
        if isinstance(capped, MultiPolygon):
            capped = max(capped.geoms, key=lambda p: p.area)
        area_m2, _ = geod.geometry_area_perimeter(capped)
        result[hosp_name] = (capped, abs(area_m2) / 1e6)

    return result


# ---------------------------------------------------------------------------
# Governorate intersection
# ---------------------------------------------------------------------------

def load_governorates(admin2_path: Path, adm1_filter: str = GAZA_ADM1_NAME):
    """Load admin-2 polygons filtered to the specified adm1 region."""
    gdf = gpd.read_file(admin2_path)
    if gdf.crs is None:
        gdf = gdf.set_crs(CRS_WGS84)
    gdf = gdf.to_crs(CRS_WGS84)
    gaza_govs = gdf[gdf["adm1_name"] == adm1_filter].copy()
    return gaza_govs.reset_index(drop=True)


def area_of_intersection_km2(polygon: Any, gov_polygon: Any) -> float:
    """Geodesic area of the intersection between a catchment and a governorate."""
    if polygon is None or polygon.is_empty:
        return 0.0
    try:
        inter = polygon.intersection(gov_polygon)
    except Exception:
        return 0.0
    if inter is None or inter.is_empty:
        return 0.0
    geod = Geod(ellps="WGS84")
    area_m2, _ = geod.geometry_area_perimeter(inter)
    return abs(area_m2) / 1e6


# ---------------------------------------------------------------------------
# Discover timeline periods from HTML filenames
# ---------------------------------------------------------------------------

def discover_html_periods(html_dir: Path) -> List[Tuple[datetime, datetime]]:
    """
    Parse *_catchmaps.html filenames to recover (period_start, period_end) pairs.
    Filename format: YYYYMMDD_YYYYMMDD_catchmaps.html
    """
    periods = []
    for p in sorted(html_dir.glob("*_catchmaps.html")):
        m = re.match(r"(\d{8})_(\d{8})_catchmaps\.html", p.name)
        if m:
            start = datetime.strptime(m.group(1), "%Y%m%d")
            end   = datetime.strptime(m.group(2), "%Y%m%d")
            periods.append((start, end))
    return periods


# ---------------------------------------------------------------------------
# Excel generation
# ---------------------------------------------------------------------------

def write_excel(rows: List[Dict], gov_names: List[str], gov_total_areas: Dict[str, float],
                output_path: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Catchment Summary"

    # --- Styles ---
    header_font   = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    header_fill   = PatternFill("solid", start_color="2F4F8F")
    govcat_fill   = PatternFill("solid", start_color="4472C4")   # blue sub-header
    govtot_fill   = PatternFill("solid", start_color="70AD47")   # green sub-header
    center_align  = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align    = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    thin          = Side(border_style="thin", color="AAAAAA")
    border        = Border(left=thin, right=thin, top=thin, bottom=thin)
    num_fmt       = '#,##0.00'

    # --- Build header rows (two rows: group + column name) ---
    # Fixed columns: Timeline Range, Hospital, Total Catchment Area
    # Then per-gov catchment columns, then per-gov total area columns

    fixed_headers = [
        ("", "Timeline Range"),
        ("", "Hospital"),
        ("", "Total Catchment\nArea (km²)"),
    ]
    catchment_headers = [(f"{g}\nCatchment (km²)", g) for g in gov_names]
    total_headers     = [(f"{g}\nTotal Area (km²)", g) for g in gov_names]

    # Row 1 (groups), Row 2 (column names) approach with merged cells for fixed cols
    all_col_defs = fixed_headers + catchment_headers + total_headers

    # Write two header rows
    # Row 1: group labels
    # Row 2: column labels
    group_row = 1
    col_row   = 2

    col = 1
    # Fixed cols: merge rows 1-2
    for (group, colname) in fixed_headers:
        cell = ws.cell(row=group_row, column=col, value=colname)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = border
        ws.merge_cells(start_row=group_row, start_column=col,
                       end_row=col_row,   end_column=col)
        col += 1

    # Catchment columns group label
    catch_start = col
    catch_end   = col + len(gov_names) - 1
    ws.cell(row=group_row, column=catch_start,
            value="Catchment Area by Governorate (km²)").font = header_font
    ws.cell(row=group_row, column=catch_start).fill  = govcat_fill
    ws.cell(row=group_row, column=catch_start).alignment = center_align
    if len(gov_names) > 1:
        ws.merge_cells(start_row=group_row, start_column=catch_start,
                       end_row=group_row,   end_column=catch_end)
    for g in gov_names:
        c = ws.cell(row=col_row, column=col, value=g)
        c.font = Font(name="Arial", bold=True, size=10)
        c.fill = govcat_fill
        c.alignment = center_align
        c.border = border
        col += 1

    # Total area group label
    tot_start = col
    tot_end   = col + len(gov_names) - 1
    ws.cell(row=group_row, column=tot_start,
            value="Total Governorate Area (km²)").font = header_font
    ws.cell(row=group_row, column=tot_start).fill  = govtot_fill
    ws.cell(row=group_row, column=tot_start).alignment = center_align
    if len(gov_names) > 1:
        ws.merge_cells(start_row=group_row, start_column=tot_start,
                       end_row=group_row,   end_column=tot_end)
    for g in gov_names:
        c = ws.cell(row=col_row, column=col, value=g)
        c.font = Font(name="Arial", bold=True, size=10)
        c.fill = govtot_fill
        c.alignment = center_align
        c.border = border
        col += 1

    # Style row 1 borders for group cells
    for cc in range(1, col):
        c = ws.cell(row=group_row, column=cc)
        c.border = border

    # --- Data rows ---
    for data_row_idx, row in enumerate(rows, start=3):
        col = 1
        # Timeline range
        c = ws.cell(row=data_row_idx, column=col, value=row["timeline_range"])
        c.font = Font(name="Arial", size=10)
        c.alignment = left_align
        c.border = border
        col += 1
        # Hospital
        c = ws.cell(row=data_row_idx, column=col, value=row["hospital"])
        c.font = Font(name="Arial", size=10)
        c.alignment = left_align
        c.border = border
        col += 1
        # Total catchment area
        c = ws.cell(row=data_row_idx, column=col, value=round(row["total_catchment_km2"], 4))
        c.font = Font(name="Arial", size=10)
        c.number_format = num_fmt
        c.alignment = center_align
        c.border = border
        col += 1
        # Per-gov catchment
        for g in gov_names:
            c = ws.cell(row=data_row_idx, column=col,
                        value=round(row["gov_catchment"].get(g, 0.0), 4))
            c.font = Font(name="Arial", size=10)
            c.number_format = num_fmt
            c.alignment = center_align
            c.border = border
            col += 1
        # Per-gov total area
        for g in gov_names:
            c = ws.cell(row=data_row_idx, column=col,
                        value=round(gov_total_areas.get(g, 0.0), 4))
            c.font = Font(name="Arial", size=10)
            c.number_format = num_fmt
            c.alignment = center_align
            c.border = border
            col += 1

    # --- Column widths ---
    ws.column_dimensions[get_column_letter(1)].width = 26   # timeline
    ws.column_dimensions[get_column_letter(2)].width = 28   # hospital
    ws.column_dimensions[get_column_letter(3)].width = 18   # total catchment
    for i in range(4, 4 + 2 * len(gov_names)):
        ws.column_dimensions[get_column_letter(i)].width = 20

    # Freeze panes below header
    ws.freeze_panes = "A3"

    # Row heights
    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 36

    wb.save(output_path)
    print(f"  Saved: {output_path}")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    print("=" * 70)
    print("Catchment Summary Excel Generator")
    print("=" * 70)

    # 1. Load hospitals & timeline
    print("\n[1] Loading hospital availability data...")
    hospitals_df, availability_timeline = read_hospitals_open_close(HOSP_PATH)
    print(f"    {len(hospitals_df)} hospitals loaded")

    # 2. Load Gaza boundary (for Voronoi clipping)
    print("\n[2] Loading Gaza boundary...")
    boundary_gdf = gpd.read_file(GAZA_GEOJSON)
    if boundary_gdf.crs is None:
        boundary_gdf = boundary_gdf.set_crs(CRS_WGS84)
    boundary_gdf = boundary_gdf.to_crs(CRS_WGS84)
    gaza_union = boundary_gdf.unary_union

    # 3. Load governorates
    print("\n[3] Loading Gaza Strip governorates from pse_admin2.geojson...")
    gov_gdf = load_governorates(ADMIN2_GEOJSON)
    gov_names = gov_gdf["adm2_name"].tolist()
    gov_total_areas = {
        row["adm2_name"]: float(row["area_sqkm"])
        for _, row in gov_gdf.iterrows()
    }
    print(f"    Governorates: {gov_names}")

    # 4. Discover timeline periods from HTML filenames
    print("\n[4] Discovering timeline periods from HTML files...")
    html_periods = discover_html_periods(HTML_DIR)
    if not html_periods:
        # Fall back: re-derive from availability timeline
        print("    No HTML files found — deriving periods from timeline data.")
        html_periods = get_all_status_change_periods(hospitals_df, availability_timeline)
    print(f"    Found {len(html_periods)} periods")

    # 5. Build Excel rows
    print("\n[5] Computing catchment × governorate intersections...")
    excel_rows = []

    for period_start, period_end in html_periods:
        timeline_range = (
            f"{period_start.strftime('%Y-%m-%d')} to {period_end.strftime('%Y-%m-%d')}"
        )
        open_hospitals = get_open_hospitals_at_date(
            hospitals_df, availability_timeline, period_start
        )
        if not open_hospitals:
            print(f"    SKIP {timeline_range}: no open hospitals")
            continue

        catchments = catchment_area_method(open_hospitals, gaza_union)

        for hosp_name, (poly, total_area_km2) in catchments.items():
            gov_catchment = {}
            for _, gov_row in gov_gdf.iterrows():
                g = gov_row["adm2_name"]
                gov_catchment[g] = area_of_intersection_km2(poly, gov_row.geometry)

            excel_rows.append({
                "timeline_range":    timeline_range,
                "hospital":          hosp_name,
                "total_catchment_km2": total_area_km2,
                "gov_catchment":     gov_catchment,
            })

        print(f"    {timeline_range}: {len(open_hospitals)} hospitals")

    # 6. Write Excel
    print(f"\n[6] Writing {len(excel_rows)} rows to Excel...")
    write_excel(excel_rows, gov_names, gov_total_areas, OUTPUT_XLSX)

    print("\nDone.")
    print("=" * 70)


if __name__ == "__main__":
    main()